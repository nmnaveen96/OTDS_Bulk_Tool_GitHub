#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════╗
║    OpenText OTDS 25.x – Bulk Users & Groups Provisioning Tool       ║
║    Version : 1.0.0   |   Requires: Python 3.8+                      ║
║    REST API: /otdsws/rest  (OTDS 24.x / 25.x compatible)            ║
╚══════════════════════════════════════════════════════════════════════╝

Usage:
  python otds_bulk_tool.py --init
  python otds_bulk_tool.py --dry-run
  python otds_bulk_tool.py --otds-url https://otds.example.com:8443
  python otds_bulk_tool.py --config my_config.json
"""

# SECTION 1 – Standard library imports
import argparse, json, logging, os, sys, time, warnings
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

# SECTION 2 – Third-party imports
try:
    import requests
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    print(f"[ERROR] Missing dependency: {exc}")
    print("        Run: pip install requests pandas openpyxl")
    sys.exit(1)

warnings.filterwarnings("ignore", message="Unverified HTTPS request")

# SECTION 3 – Defaults & branding
DEFAULT_CONFIG: Dict[str, Any] = {
    "otds_url":    "https://otds.example.com:8443",
    "admin_user":  "otadmin@otds.admin",
    "admin_pass":  "password",
    "verify_ssl":  False,
    "timeout":     30,
    "dry_run":     False,
    "log_level":   "INFO",
    "log_file":    "otds_bulk.log",
    "report_file": "otds_bulk_report.html",
    "users_input": "users.xlsx",
    "groups_input":"groups.xlsx",
}
OT_BLUE = "0063a5"; OT_BLUE_LIGHT = "d6e9f5"; OT_WHITE = "FFFFFF"

# SECTION 4 – Logging
def setup_logging(log_level: str, log_file: Optional[str]) -> logging.Logger:
    logger = logging.getLogger("OTDS_BULK")
    logger.setLevel(getattr(logging, log_level.upper(), logging.INFO))
    fmt = logging.Formatter("[%(asctime)s] %(levelname)-8s  %(message)s", datefmt="%Y-%m-%d %H:%M:%S")
    ch = logging.StreamHandler(sys.stdout); ch.setFormatter(fmt); logger.addHandler(ch)
    if log_file:
        fh = logging.FileHandler(log_file, encoding="utf-8"); fh.setFormatter(fmt); logger.addHandler(fh)
    return logger

# SECTION 5 – OTDS REST API client
class OTDSClient:
    RETRY_COUNT = 3
    RETRY_DELAY = 2

    def __init__(self, base_url, verify_ssl, timeout, dry_run, logger):
        self.base_url = base_url.rstrip("/")
        self.api_base = f"{self.base_url}/otdsws/rest"
        self.verify = verify_ssl; self.timeout = timeout
        self.dry_run = dry_run; self.log = logger
        self.ticket = None
        self.session = requests.Session(); self.session.verify = verify_ssl

    def _headers(self):
        h = {"Content-Type": "application/json", "Accept": "application/json"}
        if self.ticket: h["OTDSTicket"] = self.ticket
        return h

    def _request(self, method, endpoint, payload=None):
        url = f"{self.api_base}{endpoint}"
        for attempt in range(1, self.RETRY_COUNT + 1):
            try:
                resp = self.session.request(method, url, headers=self._headers(), json=payload, timeout=self.timeout)
                return resp.status_code, self._safe_json(resp)
            except requests.exceptions.RequestException as exc:
                self.log.warning(f"  Attempt {attempt}/{self.RETRY_COUNT} failed: {exc}")
                if attempt < self.RETRY_COUNT: time.sleep(self.RETRY_DELAY)
        return -1, {"error": "All retry attempts exhausted"}

    @staticmethod
    def _safe_json(resp):
        try: return resp.json()
        except: return resp.text or {}

    @staticmethod
    def _extract_error(data):
        if isinstance(data, dict):
            return data.get("error_description") or data.get("message") or data.get("error") or str(data)
        return str(data)

    def authenticate(self, username, password):
        if self.dry_run:
            self.log.info("  [DRY-RUN] Skipping authentication.")
            self.ticket = "DRY-RUN-TICKET"; return True
        status, data = self._request("POST", "/authentication/credentials", {"userName": username, "userPassword": password})
        if status == 200 and isinstance(data, dict):
            self.ticket = data.get("ticket") or data.get("adminTicket")
            if self.ticket: self.log.info(f"  OK Authenticated as {username}"); return True
        self.log.error(f"  ERR Authentication failed (HTTP {status}): {data}"); return False

    def list_partitions(self):
        if self.dry_run: return []
        status, data = self._request("GET", "/partitions")
        if status == 200 and isinstance(data, dict):
            return [p.get("name","") for p in data.get("partition",[]) if p.get("name")]
        self.log.warning(f"  Could not fetch partitions (HTTP {status})"); return []

    def create_user(self, partition, user):
        login = user.get("login","").strip(); user_id = f"{login}@{partition}"
        if self.dry_run:
            self.log.info(f"  [DRY-RUN] Would create user: {user_id}")
            return "dry-run", f"Would create {user_id}"
        payload = {k:v for k,v in {
            "userID": user_id, "name": login, "partition": partition,
            "oTExtraAttr0": user.get("first_name",""), "oTExtraAttr1": user.get("last_name",""),
            "mail": user.get("email",""), "description": user.get("description",""),
            "telephoneNumber": user.get("phone",""), "title": user.get("title",""),
            "department": user.get("department",""), "company": user.get("company",""),
            "oTPassword": user.get("password",""),
        }.items() if v}
        status, data = self._request("POST", "/users", payload)
        if status in (200,201): self.log.info(f"  OK Created user: {user_id}"); return "created", f"User {user_id} created"
        if status == 409: self.log.info(f"  -> Skipped (exists): {user_id}"); return "skipped", f"User {user_id} already exists"
        msg = self._extract_error(data); self.log.error(f"  ERR {user_id}: {msg}"); return "error", msg

    def create_group(self, partition, group):
        name = group.get("name","").strip(); group_id = f"{name}@{partition}"
        if self.dry_run:
            self.log.info(f"  [DRY-RUN] Would create group: {group_id}")
            return "dry-run", f"Would create {group_id}"
        payload = {k:v for k,v in {
            "groupID": group_id, "name": name, "partition": partition,
            "description": group.get("description",""), "mail": group.get("email",""),
        }.items() if v}
        status, data = self._request("POST", "/groups", payload)
        if status in (200,201): self.log.info(f"  OK Created group: {group_id}"); return "created", f"Group {group_id} created"
        if status == 409: self.log.info(f"  -> Skipped (exists): {group_id}"); return "skipped", f"Group {group_id} already exists"
        msg = self._extract_error(data); self.log.error(f"  ERR {group_id}: {msg}"); return "error", msg

    def add_member_to_group(self, group_name, partition, member_login):
        group_id = f"{group_name}@{partition}"; member_id = f"{member_login}@{partition}"
        if self.dry_run:
            self.log.info(f"  [DRY-RUN] Would add {member_id} -> {group_id}")
            return "dry-run", f"Would add {member_id} to {group_id}"
        encoded = requests.utils.quote(group_id, safe="")
        status, data = self._request("POST", f"/groups/{encoded}/members", {"memberID": [member_id]})
        if status in (200,201,204): self.log.info(f"  OK Added {member_id} -> {group_id}"); return "created", f"Added {member_id} to {group_id}"
        if status == 409: return "skipped", f"{member_id} already in {group_id}"
        msg = self._extract_error(data); self.log.error(f"  ERR {msg}"); return "error", msg

# SECTION 6 – Excel reader
def read_excel_by_partition(filepath, logger):
    result = {}
    if not os.path.exists(filepath):
        logger.warning(f"  File not found: {filepath}"); return result
    xl = pd.ExcelFile(filepath, engine="openpyxl"); sheets = xl.sheet_names
    if len(sheets) > 1:
        for sheet in sheets:
            df = pd.read_excel(filepath, sheet_name=sheet, engine="openpyxl", dtype=str)
            df = df.where(pd.notnull(df),""); df.columns=[c.strip().lower().replace(" ","_") for c in df.columns]
            result[sheet] = df.to_dict(orient="records")
    else:
        df = pd.read_excel(filepath, sheet_name=0, engine="openpyxl", dtype=str)
        df = df.where(pd.notnull(df),""); df.columns=[c.strip().lower().replace(" ","_") for c in df.columns]
        if "partition" in df.columns:
            for part, grp in df.groupby("partition"): result[str(part)] = grp.to_dict(orient="records")
        else:
            result[sheets[0]] = df.to_dict(orient="records")
    return result

# SECTION 7 – Excel template generator
def _apply_header_style(ws, headers):
    hf=PatternFill("solid",fgColor=OT_BLUE); hfont=Font(bold=True,color=OT_WHITE,name="Calibri",size=11)
    af=PatternFill("solid",fgColor=OT_BLUE_LIGHT); df=PatternFill("solid",fgColor=OT_WHITE)
    ca=Alignment(horizontal="center",vertical="center",wrap_text=True)
    la=Alignment(horizontal="left",vertical="center")
    bs=Side(style="thin",color="AAAAAA"); tb=Border(left=bs,right=bs,top=bs,bottom=bs)
    samples={"login":["jsmith","mjones"],"first_name":["John","Mary"],"last_name":["Smith","Jones"],
             "email":["jsmith@org.qa","mjones@org.qa"],"password":["P@ssword1!","P@ssword2!"],
             "title":["Senior Analyst","Manager"],"department":["IT","Finance"],
             "company":["Qatar Foundation","Qatar Foundation"],"phone":["+97430000001","+97430000002"],
             "description":["IT Staff","Finance Manager"],"groups":["IT_Admins;ECM_Users","Finance_Grp"],
             "name":["IT_Admins","Finance_Grp"],"members":["jsmith;mjones","mjones"],"partition":["Partition_A","Partition_B"]}
    for ci,h in enumerate(headers,1):
        c=ws.cell(1,ci,h); c.fill=hf; c.font=hfont; c.alignment=ca; c.border=tb
    for ri in range(2,4):
        i=ri-2; fill=af if ri%2==0 else df
        for ci,h in enumerate(headers,1):
            v=samples.get(h,["sample","sample"])[i]
            c=ws.cell(ri,ci,v); c.fill=fill; c.font=Font(name="Calibri",size=10); c.alignment=la; c.border=tb
    ws.freeze_panes="A2"; ws.row_dimensions[1].height=22
    for ci,h in enumerate(headers,1): ws.column_dimensions[get_column_letter(ci)].width=max(len(h),18)+2

def generate_templates(users_file, groups_file, logger):
    uh=["login","first_name","last_name","email","password","title","department","company","phone","description","groups"]
    gh=["name","description","email","members"]
    wb_u=openpyxl.Workbook(); wb_u.remove(wb_u.active)
    for p in ["Partition_A","Partition_B"]: _apply_header_style(wb_u.create_sheet(p),uh)
    wb_u.save(users_file); logger.info(f"  OK Users template -> {users_file}")
    wb_g=openpyxl.Workbook(); wb_g.remove(wb_g.active)
    for p in ["Partition_A","Partition_B"]: _apply_header_style(wb_g.create_sheet(p),gh)
    wb_g.save(groups_file); logger.info(f"  OK Groups template -> {groups_file}")

# SECTION 8 – HTML report generator
def generate_html_report(results, config, start_time, report_file, logger):
    end_time=datetime.now(); duration=(end_time-start_time).total_seconds(); dry_run=config.get("dry_run",False)
    total=len(results); created=sum(1 for r in results if r["status"]=="created")
    skipped=sum(1 for r in results if r["status"]=="skipped"); errors=sum(1 for r in results if r["status"]=="error")
    dry_cnt=sum(1 for r in results if r["status"]=="dry-run")
    rows=""
    for r in results:
        css={"created":"row-green","skipped":"row-yellow","error":"row-red","dry-run":"row-blue"}.get(r["status"],"")
        icon={"created":"&#10004;","skipped":"&#8594;","error":"&#10008;","dry-run":"&#9702;"}.get(r["status"],"")
        rows+=f'<tr class="{css}"><td>{r.get("partition","")}</td><td>{r.get("type","")}</td><td><code>{r.get("name","")}</code></td><td>{icon} {r.get("status","").upper()}</td><td>{r.get("detail","")}</td></tr>'
    db='<div class="dry-banner">&#9888; DRY-RUN MODE &#8211; No changes were made to OTDS</div>' if dry_run else ""
    dc=f'<div class="card card-dryrun"><div class="number">{dry_cnt}</div><div class="label">&#9702; Dry-Run</div></div>' if dry_run else ""
    html=f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"/>
<title>OTDS Bulk Provisioning Report</title>
<style>*{{box-sizing:border-box;margin:0;padding:0}}body{{font-family:'Segoe UI',Arial,sans-serif;background:#f0f4f8;color:#222}}
.header{{background:linear-gradient(135deg,#0063a5,#004a7c);color:#fff;padding:28px 40px}}
.header h1{{font-size:1.8rem;font-weight:700}}.header .sub{{font-size:.85rem;opacity:.85;margin-top:6px}}
.dry-banner{{background:#fff3cd;color:#856404;border-left:5px solid #ffc107;padding:12px 40px;font-weight:600}}
.cards{{display:flex;gap:18px;padding:28px 40px 0;flex-wrap:wrap}}
.card{{flex:1;min-width:140px;border-radius:10px;padding:20px 24px;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,.1)}}
.card .number{{font-size:2.4rem;font-weight:800}}.card .label{{font-size:.8rem;text-transform:uppercase;margin-top:4px}}
.card-total{{background:#fff;border-top:4px solid #0063a5}}.card-created{{background:#d4edda;border-top:4px solid #28a745}}
.card-skipped{{background:#fff3cd;border-top:4px solid #ffc107}}.card-error{{background:#f8d7da;border-top:4px solid #dc3545}}
.card-dryrun{{background:#d6e9f5;border-top:4px solid #17a2b8}}
.info-bar{{margin:22px 40px 0;padding:12px 18px;background:#fff;border-radius:8px;font-size:.84rem;color:#555;box-shadow:0 1px 4px rgba(0,0,0,.08);display:flex;gap:30px;flex-wrap:wrap}}
.info-bar span b{{color:#0063a5}}.table-wrap{{margin:22px 40px 40px;overflow-x:auto;border-radius:10px;box-shadow:0 2px 12px rgba(0,0,0,.1)}}
table{{width:100%;border-collapse:collapse;background:#fff;font-size:.88rem}}
thead tr{{background:#0063a5;color:#fff;text-transform:uppercase;font-size:.78rem}}
th,td{{padding:10px 14px;border-bottom:1px solid #e0e0e0}}
code{{background:#f0f4f8;padding:2px 6px;border-radius:4px;font-size:.82rem}}
.row-green{{background:#d4edda}}.row-yellow{{background:#fff3cd}}.row-red{{background:#f8d7da}}.row-blue{{background:#d6e9f5}}
tbody tr:hover{{filter:brightness(.97)}}.footer{{text-align:center;padding:18px;font-size:.78rem;color:#888;border-top:1px solid #ddd}}
.footer a{{color:#0063a5}}</style></head><body>
<div class="header"><h1>&#128193; OpenText OTDS &#8211; Bulk Provisioning Report</h1>
<div class="sub">OTDS Server: <strong>{config.get("otds_url","N/A")}</strong> &nbsp;|&nbsp; Generated: {end_time.strftime("%Y-%m-%d %H:%M:%S")} &nbsp;|&nbsp; Duration: {duration:.1f}s</div></div>
{db}
<div class="cards">
<div class="card card-total"><div class="number">{total}</div><div class="label">Total Operations</div></div>
<div class="card card-created"><div class="number">{created}</div><div class="label">&#10004; Created</div></div>
<div class="card card-skipped"><div class="number">{skipped}</div><div class="label">&#8594; Skipped</div></div>
<div class="card card-error"><div class="number">{errors}</div><div class="label">&#10008; Errors</div></div>{dc}</div>
<div class="info-bar">
<span><b>Admin:</b> {config.get("admin_user","")}</span><span><b>Users:</b> {config.get("users_input","")}</span>
<span><b>Groups:</b> {config.get("groups_input","")}</span><span><b>SSL:</b> {config.get("verify_ssl","")}</span>
<span><b>Mode:</b> {"&#128993; DRY-RUN" if dry_run else "&#128994; LIVE"}</span></div>
<div class="table-wrap"><table><thead><tr><th>Partition</th><th>Type</th><th>Name</th><th>Status</th><th>Detail</th></tr></thead>
<tbody>{rows if rows else '<tr><td colspan="5" style="text-align:center;color:#999">No operations recorded.</td></tr>'}</tbody></table></div>
<div class="footer">OpenText OTDS Bulk Provisioning Tool v1.0.0 &nbsp;|&nbsp;
<a href="https://developer.opentext.com/ce/products/opentext-directory-services">OTDS Developer Docs</a></div>
</body></html>"""
    try:
        with open(report_file,"w",encoding="utf-8") as f: f.write(html)
        logger.info(f"  OK HTML report saved -> {report_file}")
    except Exception as exc: logger.error(f"  ERR Could not write report: {exc}")

# SECTION 9 – Config loader
def load_config(config_path, logger):
    cfg=dict(DEFAULT_CONFIG)
    if os.path.exists(config_path):
        try:
            with open(config_path,"r",encoding="utf-8") as f: user_cfg=json.load(f)
            cfg.update(user_cfg); logger.info(f"  OK Loaded config: {config_path}")
        except Exception as exc: logger.warning(f"  Could not parse {config_path}: {exc}")
    else: logger.info(f"  Config not found ({config_path}), using defaults.")
    return cfg

def save_config_template(path, logger):
    t=dict(DEFAULT_CONFIG); t["_comment"]="Set dry_run=true to preview without making changes."
    with open(path,"w",encoding="utf-8") as f: json.dump(t,f,indent=2)
    logger.info(f"  OK Config template -> {path}")

# SECTION 10 – Provisioning engine
def provision(config, logger):
    results=[]; dry_run=config.get("dry_run",False)
    client=OTDSClient(config["otds_url"],config.get("verify_ssl",False),config.get("timeout",30),dry_run,logger)
    logger.info("\n"+"="*65)
    logger.info(" OTDS BULK PROVISIONING  –  "+("DRY-RUN MODE" if dry_run else "LIVE MODE"))
    logger.info("="*65)
    logger.info(f"\n[1/4] Authenticating -> {config['otds_url']}")
    if not client.authenticate(config["admin_user"],config["admin_pass"]):
        logger.error("  Authentication failed. Aborting."); sys.exit(1)
    logger.info("\n[2/4] Fetching existing OTDS partitions ...")
    existing=client.list_partitions()
    logger.info(f"  Found {len(existing)} partition(s): {', '.join(existing)}" if existing else "  (Unavailable in dry-run mode)")
    logger.info("\n[3/4] Reading input files ...")
    gbp=read_excel_by_partition(config.get("groups_input","groups.xlsx"),logger)
    ubp=read_excel_by_partition(config.get("users_input","users.xlsx"),logger)
    all_p=sorted(set(list(gbp.keys())+list(ubp.keys())))
    if not all_p: logger.warning("  No partitions found. Nothing to do."); return results
    logger.info(f"  Partitions to process: {', '.join(all_p)}")
    logger.info("\n[4/4] Provisioning ...")
    for partition in all_p:
        logger.info(f"\n{'─'*60}\n  PARTITION: {partition}\n{'─'*60}")
        groups=gbp.get(partition,[])
        if groups: logger.info(f"\n  Creating {len(groups)} group(s) ...")
        for grp in groups:
            name=grp.get("name","").strip()
            if not name: continue
            s,d=client.create_group(partition,grp)
            results.append({"partition":partition,"type":"Group","name":name,"status":s,"detail":d})
        users=ubp.get(partition,[])
        if users: logger.info(f"\n  Creating {len(users)} user(s) ...")
        ugm={}
        for usr in users:
            login=usr.get("login","").strip()
            if not login: continue
            s,d=client.create_user(partition,usr)
            results.append({"partition":partition,"type":"User","name":login,"status":s,"detail":d})
            rg=usr.get("groups","").strip()
            if rg: ugm[login]=[g.strip() for g in rg.split(";") if g.strip()]
        if ugm: logger.info("\n  Assigning user->group memberships (users sheet) ...")
        for login,gnames in ugm.items():
            for gn in gnames:
                s,d=client.add_member_to_group(gn,partition,login)
                results.append({"partition":partition,"type":"Membership","name":f"{login} -> {gn}","status":s,"detail":d})
        gwm=[g for g in groups if g.get("members","").strip()]
        if gwm: logger.info("\n  Assigning members (groups sheet) ...")
        for grp in gwm:
            gn=grp.get("name","").strip(); rm=grp.get("members","").strip()
            if not gn or not rm: continue
            for m in [x.strip() for x in rm.split(";") if x.strip()]:
                s,d=client.add_member_to_group(gn,partition,m)
                results.append({"partition":partition,"type":"Membership","name":f"{m} -> {gn}","status":s,"detail":d})
    return results

# SECTION 11 – CLI parser
def build_parser():
    p=argparse.ArgumentParser(prog="otds_bulk_tool.py",
        description="OpenText OTDS 25.x – Bulk Users & Groups Provisioning Tool",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Examples:\n  python otds_bulk_tool.py --init\n  python otds_bulk_tool.py --dry-run\n  python otds_bulk_tool.py --otds-url https://otds.qf.org.qa:8443")
    p.add_argument("--init",action="store_true",help="Generate config + Excel templates and exit")
    p.add_argument("--config",default="config.json",metavar="FILE",help="Config JSON file (default: config.json)")
    p.add_argument("--otds-url",dest="otds_url",metavar="URL",help="Override OTDS base URL")
    p.add_argument("--users",dest="users_input",metavar="XLSX",help="Override users Excel path")
    p.add_argument("--groups",dest="groups_input",metavar="XLSX",help="Override groups Excel path")
    p.add_argument("--dry-run",dest="dry_run",action="store_true",help="Preview without API calls")
    p.add_argument("--log-level",dest="log_level",choices=["DEBUG","INFO","WARNING","ERROR"],default=None)
    p.add_argument("--report",dest="report_file",metavar="HTML",help="Override HTML report path")
    return p

# SECTION 12 – Entry point
def main():
    parser=build_parser(); args=parser.parse_args()
    logger=setup_logging("INFO",None)
    if args.init:
        logger.info("\n"+"="*60+"\n  OTDS Bulk Tool – Initializing templates\n"+"="*60)
        save_config_template("config.json",logger)
        generate_templates("users.xlsx","groups.xlsx",logger)
        logger.info("\n  Done! Next steps:")
        logger.info("  1. Edit config.json with your OTDS URL + credentials")
        logger.info("  2. Populate users.xlsx and groups.xlsx (sheet = partition)")
        logger.info("  3. Run: python otds_bulk_tool.py --dry-run")
        logger.info("  4. Run: python otds_bulk_tool.py"); return
    config=load_config(args.config,logger)
    if args.otds_url:    config["otds_url"]=args.otds_url
    if args.users_input: config["users_input"]=args.users_input
    if args.groups_input:config["groups_input"]=args.groups_input
    if args.dry_run:     config["dry_run"]=True
    if args.log_level:   config["log_level"]=args.log_level
    if args.report_file: config["report_file"]=args.report_file
    for h in logger.handlers[:]: logger.removeHandler(h)
    logger=setup_logging(config.get("log_level","INFO"),config.get("log_file"))
    start=datetime.now(); results=provision(config,logger)
    total=len(results); created=sum(1 for r in results if r["status"]=="created")
    skipped=sum(1 for r in results if r["status"]=="skipped"); errors=sum(1 for r in results if r["status"]=="error")
    dry_cnt=sum(1 for r in results if r["status"]=="dry-run")
    logger.info("\n"+"="*65+"\n  PROVISIONING COMPLETE")
    logger.info(f"  Total: {total} | Created: {created} | Skipped: {skipped} | Errors: {errors}"+(f" | Dry-Run: {dry_cnt}" if config.get("dry_run") else ""))
    logger.info("="*65)
    generate_html_report(results,config,start,config.get("report_file","otds_bulk_report.html"),logger)

if __name__ == "__main__":
    main()
